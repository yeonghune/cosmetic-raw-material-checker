from openpyxl import Workbook, load_workbook
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from app.models import IngredientRow, DiffType
from app.ui.styles import AppColors
from app.utils.diff_logic import generate_diff_report

FIXED_HEADER = ("RM", "% RM/FP", "INCI", "% INCI/RM")

def download_template_file(output_path: str | Path = "다운로드/output.xlsx") -> Path:
    """빈 템플릿 엑셀 파일을 생성합니다."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Table1"
    
    ws2 = wb.create_sheet(title="Table2")
    
    for ws in [ws1, ws2]:
        for col_idx, header in enumerate(FIXED_HEADER, start=1):
            ws.cell(row=1, column=col_idx).value = header

    output = Path(output_path)
    if output.parent:
        output.parent.mkdir(parents=True, exist_ok=True)
        
    wb.save(output)
    return output


def load_data_from_excel(file_path: str, sheet_name: str) -> list[IngredientRow]:
    """엑셀 파일에서 원시 데이터를 읽어 IngredientRow 리스트로 변환합니다."""
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
        
    sheet = wb[sheet_name]
    raw_data = []
    
    # Fill-down을 위한 변수
    current_rm = ""
    current_rm_pct = ""

    # 2행(헤더 다음)부터 읽기
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 안전한 접근 (인덱스 초과 방지)
        rm_val = str(row[0]) if len(row) > 0 and row[0] is not None else None
        rm_pct = str(row[1]) if len(row) > 1 and row[1] is not None else ""
        inci = str(row[2]) if len(row) > 2 and row[2] is not None else ""
        inci_pct = str(row[3]) if len(row) > 3 and row[3] is not None else ""

        # 새 RM이 나오면 업데이트, 없으면 이전 RM 사용 (Fill-down 로직)
        if rm_val:
            current_rm = rm_val
            current_rm_pct = rm_pct
        
        # 유효한 RM 그룹 내에 있다면 데이터 추가
        if current_rm:
            raw_data.append(IngredientRow(
                rm_name=current_rm,
                rm_percent=current_rm_pct,
                inci_name=inci,
                inci_percent=inci_pct
            ))
            
    return raw_data

def export_to_excel(output_path: str, data1: list[IngredientRow], data2: list[IngredientRow]):
    """
    두 테이블의 데이터를 엑셀로 내보냅니다.
    - 숫자 변환 (String -> Float)
    - 셀 병합 (RM 단위)
    - 스타일 적용 (Diff Report 기반)
    """
    
    # 스타일 정의
    RED_FONT = Font(color="FF0000")
    RED_BG_FILL = PatternFill(start_color="FFC8C8", end_color="FFC8C8", fill_type="solid")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

    wb = Workbook()

    # Diff Report 생성 (스타일 적용을 위해)
    diff1 = generate_diff_report(data1, data2)
    diff2 = generate_diff_report(data2, data1)

    # -------------------------------------------------------------
    # Helper: 시트 작성 및 스타일링
    # -------------------------------------------------------------
    def _write_sheet(ws, dataset_list, diff_reports_list, start_col_list):
        from openpyxl.utils import get_column_letter

        # 1. Header 작성 & Column Width 설정
        headers = []
        for _ in dataset_list:
            headers.extend(FIXED_HEADER)
            
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.alignment = CENTER_ALIGN
            cell.font = Font(bold=True)
            
        # 컬럼 너비 조정 (RM, INCI는 넓게)
        for dataset_idx, _ in enumerate(dataset_list):
            start_col = start_col_list[dataset_idx]
            
            # RM (50)
            ws.column_dimensions[get_column_letter(start_col + 0)].width = 50
            # % RM/FP (15)
            ws.column_dimensions[get_column_letter(start_col + 1)].width = 15
            # INCI (50)
            ws.column_dimensions[get_column_letter(start_col + 2)].width = 50
            # % INCI/RM (15)
            ws.column_dimensions[get_column_letter(start_col + 3)].width = 15

        # 2. Data 작성 및 병합
        max_rows = max((len(d) for d in dataset_list), default=0)
        
        for dataset_idx, data in enumerate(dataset_list):
            start_col = start_col_list[dataset_idx]
            
            # 병합을 위한 상태 변수
            merge_start_row = 2
            if not data:
                continue
                
            prev_rm = data[0].rm_name

            for i, item in enumerate(data):
                row_idx = i + 2
                
                # 값 쓰기 (숫자 변환 포함)
                ws.cell(row=row_idx, column=start_col + 0, value=item.rm_name).alignment = CENTER_ALIGN
                ws.cell(row=row_idx, column=start_col + 1, value=_try_float(item.rm_percent)).alignment = CENTER_ALIGN
                ws.cell(row=row_idx, column=start_col + 2, value=item.inci_name).alignment = CENTER_ALIGN
                ws.cell(row=row_idx, column=start_col + 3, value=_try_float(item.inci_percent)).alignment = CENTER_ALIGN

                # 병합 로직 (RM 이름 기준)
                if item.rm_name != prev_rm:
                    # 이전 그룹 병합 (행 개수가 1개 이상일 때만)
                    if row_idx - 1 > merge_start_row:
                        ws.merge_cells(start_row=merge_start_row, start_column=start_col, end_row=row_idx-1, end_column=start_col)     # RM
                        ws.merge_cells(start_row=merge_start_row, start_column=start_col+1, end_row=row_idx-1, end_column=start_col+1) # % RM
                    
                    # 상태 업데이트
                    prev_rm = item.rm_name
                    merge_start_row = row_idx
            
            # 마지막 그룹 병합
            if (len(data) + 1) > merge_start_row:
                end_row = len(data) + 1
                ws.merge_cells(start_row=merge_start_row, start_column=start_col, end_row=end_row, end_column=start_col)
                ws.merge_cells(start_row=merge_start_row, start_column=start_col+1, end_row=end_row, end_column=start_col+1)

        # 3. 스타일(Diff) 적용
        for dataset_idx, diffs in enumerate(diff_reports_list):
            start_col = start_col_list[dataset_idx]
            
            for diff in diffs:
                target_row = diff.row + 2  # 0-indexed -> 1-indexed + Header(1)
                target_col = diff.col + start_col # 0-indexed -> 1-based start_col
                
                cell = ws.cell(row=target_row, column=target_col)
                
                if diff.diff_type == DiffType.CONTENT_MISMATCH:
                    cell.font = RED_FONT
                elif diff.diff_type in (DiffType.MISSING_ROW, DiffType.MISSING_INCI):
                    cell.fill = RED_BG_FILL

    # -------------------------------------------------------------
    # Sheet 생성 및 실행
    # -------------------------------------------------------------
    
    # Sheet 1: Result (Combined)
    ws_combined = wb.active
    ws_combined.title = "Result"
    _write_sheet(ws_combined, [data1, data2], [diff1, diff2], [1, 5])
    
    # Sheet 2: Table1
    ws_t1 = wb.create_sheet(title="Table1")
    _write_sheet(ws_t1, [data1], [diff1], [1])
    
    # Sheet 3: Table2
    ws_t2 = wb.create_sheet(title="Table2")
    _write_sheet(ws_t2, [data2], [diff2], [1])

    wb.save(output_path)
    return Path(output_path)


def _try_float(value: str):
    """문자열을 가능한 경우 float로 변환합니다."""
    if not value:
        return ""
    try:
        return float(value)
    except ValueError:
        return value


def export_comparison_table(table, file_path: str):
    """
    QTableWidget의 내용을 엑셀로 내보냅니다.
    - 배경색이 노란색(독자 성분/불일치)인 경우 셀 스타일 적용
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Result"
    
    # Headers
    headers = ["A열 성분", "B열 성분"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Yellow Fill Definition
    YELLOW_FILL = PatternFill(start_color=AppColors.DIFF_BG_YELLOW_HEX, end_color=AppColors.DIFF_BG_YELLOW_HEX, fill_type="solid")
    
    # Logic to detect if a cell is highlighted
    target_qcolor = AppColors.DIFF_BG_YELLOW
    target_rgb = (target_qcolor.red(), target_qcolor.green(), target_qcolor.blue())
    
    rows = table.rowCount()
    cols = table.columnCount()
    
    for r in range(rows):
        for c in range(cols):
            item = table.item(r, c)
            text = item.text() if item else ""
            cell = ws.cell(row=r+2, column=c+1, value=text)
            
            # Check Background Color
            if item:
                bg_color = item.background().color()
                if bg_color.isValid():
                    current_rgb = (bg_color.red(), bg_color.green(), bg_color.blue())
                    if current_rgb == target_rgb:
                        cell.fill = YELLOW_FILL
                 
    # Resize columns
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    
    wb.save(file_path)
