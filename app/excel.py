from openpyxl import Workbook, load_workbook
from pathlib import Path

from PyQt5.QtWidgets import QTableWidgetItem, QHeaderView
from PyQt5.QtGui import QColor, QBrush

from app.models import IngredientRow

# ---------------------------------------------------------
# 상수 및 전역 설정
# ---------------------------------------------------------
FIXED_HEADER = ("RM", "% RM/FP", "INCI", "% INCI/RM")

COLOR_RED = QColor(255, 0, 0)
COLOR_WHITE = QColor(255, 255, 255)
COLOR_BLACK = QColor(0, 0, 0)
COLOR_BG_RED = QColor(255, 200, 200)

# ---------------------------------------------------------
# 핵심 로직 함수 (엑셀 I/O 및 데이터 처리)
# ---------------------------------------------------------

def download_template_file(output_path: str | Path = "다운로드/output.xlsx") -> Path:
    """빈 템플릿 엑셀 파일을 생성합니다."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Table1"
    
    ws2 = wb.create_sheet(title="Table2")
    
    for ws in [ws1, ws2]:
        for col_idx, header in enumerate(FIXED_HEADER, start=1):
            ws.cell(row=1, column=col_idx).value = header

    output = Path(output_path)
    if output.parent:
        output.parent.mkdir(parents=True, exist_ok=True)
        
    wb.save(output)
    return output


def load_data_from_excel(file_path: str, sheet_name: str) -> list[IngredientRow]:
    """엑셀 파일에서 원시 데이터를 읽어 IngredientRow 리스트로 변환합니다."""
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
        
    sheet = wb[sheet_name]
    raw_data = []
    
    # Fill-down을 위한 변수
    current_rm = ""
    current_rm_pct = ""

    # 2행(헤더 다음)부터 읽기
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 안전한 접근 (인덱스 초과 방지)
        rm_val = str(row[0]) if len(row) > 0 and row[0] is not None else None
        rm_pct = str(row[1]) if len(row) > 1 and row[1] is not None else ""
        inci = str(row[2]) if len(row) > 2 and row[2] is not None else ""
        inci_pct = str(row[3]) if len(row) > 3 and row[3] is not None else ""

        # 새 RM이 나오면 업데이트, 없으면 이전 RM 사용 (Fill-down 로직)
        if rm_val:
            current_rm = rm_val
            current_rm_pct = rm_pct
        
        # 유효한 RM 그룹 내에 있다면 데이터 추가
        if current_rm:
            raw_data.append(IngredientRow(
                rm_name=current_rm,
                rm_percent=current_rm_pct,
                inci_name=inci,
                inci_percent=inci_pct
            ))
            
    return raw_data


def extract_data_from_table(table) -> list[IngredientRow]:
    """QTableWidget의 내용을 읽어 IngredientRow 리스트로 변환합니다."""
    rows = table.rowCount()
    raw_data = []
    
    for r in range(rows):
        item_rm = table.item(r, 0)
        item_rm_pct = table.item(r, 1)
        item_inci = table.item(r, 2)
        item_inci_pct = table.item(r, 3)

        raw_data.append(IngredientRow(
            rm_name=item_rm.text() if item_rm else "",
            rm_percent=item_rm_pct.text() if item_rm_pct else "",
            inci_name=item_inci.text() if item_inci else "",
            inci_percent=item_inci_pct.text() if item_inci_pct else ""
        ))
    return raw_data


# ---------------------------------------------------------
# UI 렌더링 함수 (QTableWidget 조작)
# ---------------------------------------------------------

def setup_table_header(table):
    """테이블의 헤더와 컬럼 설정을 초기화합니다."""
    table.setColumnCount(len(FIXED_HEADER))
    table.setHorizontalHeaderLabels(FIXED_HEADER)
    
    header = table.horizontalHeader()
    header.setSectionResizeMode(QHeaderView.Interactive)
    header.setSectionResizeMode(2, QHeaderView.Stretch)


def render_table(table, data_list: list[IngredientRow]):
    """데이터 리스트를 테이블에 그리고, 자동 병합을 수행합니다."""
    # 1. 정렬 (RM 이름 -> INCI 이름 순)
    if data_list:
        data_list.sort(key=lambda x: (x.rm_name, x.inci_name))

    # 2. 초기화
    table.clearContents()
    table.clearSpans()  # [신규] 기존 병합 정보 초기화 (필수)
    table.setRowCount(len(data_list))
    
    if not data_list:
        return

    # 3. 데이터 쓰기 및 병합
    merge_start_idx = 0
    prev_rm = data_list[0].rm_name

    for i, item in enumerate(data_list):
        # 아이템 생성 및 삽입
        table.setItem(i, 0, QTableWidgetItem(item.rm_name))
        table.setItem(i, 1, QTableWidgetItem(item.rm_percent))
        table.setItem(i, 2, QTableWidgetItem(item.inci_name))
        table.setItem(i, 3, QTableWidgetItem(item.inci_percent))

        # RM이 바뀌면 직전 그룹 병합 처리
        if item.rm_name != prev_rm:
            _apply_merge(table, merge_start_idx, i - merge_start_idx)
            
            # 상태 업데이트
            prev_rm = item.rm_name
            merge_start_idx = i

    # 마지막 그룹 병합 처리
    _apply_merge(table, merge_start_idx, len(data_list) - merge_start_idx)


def _apply_merge(table, start_row, span_count):
    """내부 헬퍼: 조건에 맞으면 셀 병합 수행"""
    if span_count > 1:
        table.setSpan(start_row, 0, span_count, 1) # RM 컬럼
        table.setSpan(start_row, 1, span_count, 1) # % RM/FP 컬럼


# ---------------------------------------------------------
# 메인 기능 함수 (외부 호출용)
# ---------------------------------------------------------

def make_table(table, file_path, sheet_name):
    """엑셀 파일을 읽어 테이블을 구성합니다."""
    setup_table_header(table)
    data = load_data_from_excel(file_path, sheet_name)
    render_table(table, data)


def re_sort_table(table):
    """현재 테이블 내용을 읽어서 다시 정렬하고 그립니다."""
    current_data = extract_data_from_table(table)
    render_table(table, current_data)


# ---------------------------------------------------------
# 비교 및 스타일링 로직 (Logic Separation)
# ---------------------------------------------------------

from app.models import DiffType, DiffItem

def generate_diff_report(source_data: list[IngredientRow], ref_data: list[IngredientRow]) -> list[DiffItem]:
    """
    Source(내꺼) 기준으로 Ref(상대방)와 비교하여 스타일링(Diff) 정보를 생성합니다.
    """
    diffs = []
    
    # 데이터 구조화
    struct_source = _parse_structured_data_from_list(source_data)
    struct_ref = _parse_structured_data_from_list(ref_data)

    # 비교 루프
    for rm_name, rm_info in struct_source.items():
        # Case 1.3: 내 RM이 상대방에게 아예 없음 -> 전체 행 배경 빨강
        if rm_name not in struct_ref:
            for r in rm_info["rows"]:
                for c in range(4): # 0~3 컬럼 전체
                    diffs.append(DiffItem(r, c, DiffType.MISSING_ROW))
            continue

        ref_rm = struct_ref[rm_name]

        # Case 1.1: RM 함량이 다름 -> 첫 번째 행의 % 컬럼 글자 빨강
        if rm_info["percent"] != ref_rm["percent"]:
            first_row = rm_info["rows"][0]
            diffs.append(DiffItem(first_row, 1, DiffType.CONTENT_MISMATCH))
            
        # INCI 레벨 비교
        for inci_name, inci_info in rm_info["incis"].items():
            # Case 1.4: 내 INCI가 상대방 RM 안에 없음 -> 부분 배경 빨강
            if inci_name not in ref_rm["incis"]:
                r = inci_info["row"]
                diffs.append(DiffItem(r, 2, DiffType.MISSING_INCI))
                diffs.append(DiffItem(r, 3, DiffType.MISSING_INCI))
                continue
            
            # Case 1.2: INCI 함량이 다름 -> 글자 빨강
            ref_inci = ref_rm["incis"][inci_name]
            if inci_info["percent"] != ref_inci["percent"]:
                diffs.append(DiffItem(inci_info["row"], 3, DiffType.CONTENT_MISMATCH))
                
    return diffs


# [Legacy Support] 기존 UI 직접 참조 코드는 main.py에서 widget 메서드로 대체될 예정이므로 삭제하거나 수정.
# 현재는 main.py와의 호환을 위해 유지하되, 내부에서 위 함수를 호출하도록 변경.

def compare_tables_and_apply(table1, table2):
    """
    [Deprecated Style] 두 테이블을 비교하고 스타일을 즉시 적용합니다.
    Refactored to use generate_diff_report internally.
    """
    data1 = extract_data_from_table(table1)
    data2 = extract_data_from_table(table2)

    # 1. 초기화 (위젯 메서드 호출 권장하나, 여기서는 직접 접근)
    # (실제 스타일링 초기화는 Widget의 apply_diff_report에서 처리하는 것이 좋음)
    # 여기서는 호환성을 위해 아래 로직을 유지하지 않고 Main에서 처리하도록 위임하는게 맞으나,
    # 단계적 리팩토링을 위해 bridge 역할을 수행.
    pass # Main에서 apply_diff를 호출하도록 변경할 것임.


def _parse_structured_data_from_list(data_list: list[IngredientRow]):
    """IngredientRow 리스트를 딕셔너리 구조로 변환"""
    data = {}
    for i, row in enumerate(data_list):
        rm_name = row.rm_name
        
        if rm_name not in data:
            data[rm_name] = {
                "percent": row.rm_percent,
                "rows": [], # Row Index
                "incis": {}
            }
        
        data[rm_name]["rows"].append(i)
        
        if row.inci_name:
            data[rm_name]["incis"][row.inci_name] = {
                "percent": row.inci_percent,
                "row": i
            }
    return data


    return data


def export_to_excel(output_path: str, data1: list[IngredientRow], data2: list[IngredientRow]):
    """
    두 테이블의 데이터를 엑셀로 내보냅니다.
    - 숫자 변환 (String -> Float)
    - 셀 병합 (RM 단위)
    - 스타일 적용 (Diff Report 기반)
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # 스타일 정의
    RED_FONT = Font(color="FF0000")
    RED_BG_FILL = PatternFill(start_color="FFC8C8", end_color="FFC8C8", fill_type="solid")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

    wb = Workbook()

    # Diff Report 생성 (스타일 적용을 위해)
    diff1 = generate_diff_report(data1, data2)
    diff2 = generate_diff_report(data2, data1)

    # -------------------------------------------------------------
    # Helper: 시트 작성 및 스타일링
    # -------------------------------------------------------------
    def _write_sheet(ws, dataset_list, diff_reports_list, start_col_list):
        """
        ws: 대상 워크시트
        dataset_list: [data1] 또는 [data1, data2] (사이드 바이 사이드용)
        diff_reports_list: [diff1] 또는 [diff1, diff2]
        start_col_list: [1] 또는 [1, 5] (각 데이터셋의 시작 컬럼 인덱스)
        """
        from openpyxl.utils import get_column_letter

        # 1. Header 작성 & Column Width 설정
        headers = []
        for _ in dataset_list:
            headers.extend(FIXED_HEADER)
            
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.alignment = CENTER_ALIGN
            cell.font = Font(bold=True)
            
        # 컬럼 너비 조정 (RM, INCI는 넓게)
        for dataset_idx, _ in enumerate(dataset_list):
            start_col = start_col_list[dataset_idx]
            
            # RM (50)
            ws.column_dimensions[get_column_letter(start_col + 0)].width = 50
            # % RM/FP (15)
            ws.column_dimensions[get_column_letter(start_col + 1)].width = 15
            # INCI (50)
            ws.column_dimensions[get_column_letter(start_col + 2)].width = 50
            # % INCI/RM (15)
            ws.column_dimensions[get_column_letter(start_col + 3)].width = 15

        # 2. Data 작성 및 병합
        max_rows = max((len(d) for d in dataset_list), default=0)
        
        for dataset_idx, data in enumerate(dataset_list):
            start_col = start_col_list[dataset_idx]
            
            # 병합을 위한 상태 변수
            merge_start_row = 2
            if not data:
                continue
                
            prev_rm = data[0].rm_name

            for i, item in enumerate(data):
                row_idx = i + 2
                
                # 값 쓰기 (숫자 변환 포함)
                ws.cell(row=row_idx, column=start_col + 0, value=item.rm_name).alignment = CENTER_ALIGN
                ws.cell(row=row_idx, column=start_col + 1, value=_try_float(item.rm_percent)).alignment = CENTER_ALIGN
                ws.cell(row=row_idx, column=start_col + 2, value=item.inci_name).alignment = CENTER_ALIGN
                ws.cell(row=row_idx, column=start_col + 3, value=_try_float(item.inci_percent)).alignment = CENTER_ALIGN

                # 병합 로직 (RM 이름 기준)
                if item.rm_name != prev_rm:
                    # 이전 그룹 병합 (행 개수가 1개 이상일 때만)
                    if row_idx - 1 > merge_start_row:
                        ws.merge_cells(start_row=merge_start_row, start_column=start_col, end_row=row_idx-1, end_column=start_col)     # RM
                        ws.merge_cells(start_row=merge_start_row, start_column=start_col+1, end_row=row_idx-1, end_column=start_col+1) # % RM
                    
                    # 상태 업데이트
                    prev_rm = item.rm_name
                    merge_start_row = row_idx
            
            # 마지막 그룹 병합
            if (len(data) + 1) > merge_start_row:
                end_row = len(data) + 1
                ws.merge_cells(start_row=merge_start_row, start_column=start_col, end_row=end_row, end_column=start_col)
                ws.merge_cells(start_row=merge_start_row, start_column=start_col+1, end_row=end_row, end_column=start_col+1)

        # 3. 스타일(Diff) 적용
        # DiffItem의 row, col은 0-based 인덱스이므로 엑셀(1-based + Header 1행)에 맞춰 보정 필요
        for dataset_idx, diffs in enumerate(diff_reports_list):
            start_col = start_col_list[dataset_idx]
            
            for diff in diffs:
                target_row = diff.row + 2  # 0-indexed -> 1-indexed + Header(1)
                target_col = diff.col + start_col # 0-indexed -> 1-based start_col
                
                cell = ws.cell(row=target_row, column=target_col)
                
                if diff.diff_type == DiffType.CONTENT_MISMATCH:
                    cell.font = RED_FONT
                elif diff.diff_type in (DiffType.MISSING_ROW, DiffType.MISSING_INCI):
                    cell.fill = RED_BG_FILL

    # -------------------------------------------------------------
    # Sheet 생성 및 실행
    # -------------------------------------------------------------
    
    # Sheet 1: Result (Combined)
    ws_combined = wb.active
    ws_combined.title = "Result"
    _write_sheet(ws_combined, [data1, data2], [diff1, diff2], [1, 5])
    
    # Sheet 2: Table1
    ws_t1 = wb.create_sheet(title="Table1")
    _write_sheet(ws_t1, [data1], [diff1], [1])
    
    # Sheet 3: Table2
    ws_t2 = wb.create_sheet(title="Table2")
    _write_sheet(ws_t2, [data2], [diff2], [1])

    wb.save(output_path)
    return Path(output_path)


def _try_float(value: str):
    """문자열을 가능한 경우 float로 변환합니다."""
    if not value:
        return ""
    try:
        return float(value)
    except ValueError:
        return value


